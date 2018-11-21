#!/usr/bin/python
# -*- coding: utf-8 -*-

import sys
import os
import traceback
import struct
import codecs
import re
import time
from openpyxl import load_workbook
from openpyxl.styles.colors import ColorDescriptor
from openpyxl.cell.read_only import EmptyCell
import json
from collections import OrderedDict
import random
import shutil

reload(sys)
sys.setdefaultencoding("utf-8")

# struct_info = {u"vec3":[u"Vector3",["x","y","z"],["float","float","float"]],\
#     u"p_reward_view":[u"PRewardView",["type","value"], ["int","int"]],\
#     u"p_reward_view2":[u"PRewardView2",["type","typeID","num"], ["int","int","int"]], \
#     u"p_reward_view3":[u"PRewardView3",["type","typeID","num","level","rank"], ["int","int","int","int","int"]], \
#     u"p_reward":[u"PReward",["coin", "roleExp", "gold"],["int","int","int"]],\
#     u"diff_info":[u"DiffInfo",["level", "fightForce", "mod"],["int", "int", "float"]],\
#     u"position":[u"Position",["x","y","z","o"],["float","float","float","float"]],\
#     u"spell_sound_effect":[u"SpellSoundEffect",["delay","sound","loop"],["float", "string", "bool"]]}

debug_info = False
# input("请选择文件：")

##设置当前的导出目录
curDir = os.path.dirname(__file__)
outputDir = os.path.join(curDir, "UnityProject\Assets\Code\\")
dataDir = os.path.join(curDir, "UnityProject\Assets\Build\config\\")
structPath = os.path.join(curDir, "struct.json")

struct_info = json.loads(file(structPath).read(), object_pairs_hook=OrderedDict)

print "OutputDir:", outputDir
print "configDataDir:", dataDir
if not os.path.exists(outputDir):
    os.makedirs(outputDir)
if not os.path.exists(dataDir):
    os.makedirs(dataDir)

shutil.copyfile(os.path.join(curDir, "src\Entity.cs"), os.path.join(curDir, "UnityProject\Assets\Code\Entity.cs"))
shutil.copyfile(os.path.join(curDir, "src\Main.cs.meta"), os.path.join(curDir, "UnityProject\Assets\Main.cs.meta"))
shutil.copyfile(os.path.join(curDir, "src\Main.cs"), os.path.join(curDir, "UnityProject\Assets\Main.cs"))
shutil.copyfile(os.path.join(curDir, "src\Main.unity"), os.path.join(curDir, "UnityProject\Assets\Main.unity"))

structFormat = """using System;
using System.Collections.Generic;
using UnityEngine;
using System.IO;
[Serializable]
public class @@className
{
@@fields
}"""

classFormat = u"""using System;
using System.Collections.Generic;
using UnityEngine;
using System.IO;

/// <summary>
/// @@classComment
/// </summary>
/// 
[Serializable]
public partial class @@className : Entity
{
@@fields

    public override void Decode(BinaryReader br){
@@funcBody
    }
}"""

fieldFormat = u"""    /// <summary>
    /// @@fieldComment
    /// </summary>
    public @@fieldType @@fieldName;
"""

decodeStructFunc = """public static object DecodeStruct(BinaryReader br, Type t){
@@funcBody
        return null;
    }
"""

configUtil = u"""using System;
using System.IO;
using UnityEngine;
using System.Reflection;
using System.Collections.Generic;

public class ConfigUtil{

    private static Type intType = typeof(int);
    private static Type int8Type = typeof(byte);
    private static Type int16Type = typeof(short);
    private static Type boolType = typeof(bool);
    private static Type floatType = typeof(float);
    private static Type stringType = typeof(string);
    private static Type longType = typeof(long);

    public static byte[] strBuffer = new byte[10000];

    @@DecodeStruct

    public static object DecodeArray(BinaryReader br, Type t){
        Type elem = t.GetElementType();
        if(elem == intType){
            return DecodeIntArray(br);
        }else if(elem == int8Type){
            return DecodeInt8Array(br);
        }else if(elem == int16Type){
            return DecodeInt16Array(br);
        }else if(elem == floatType){
            return DecodeFloatArray(br);
        }else if(elem == stringType){
            return DecodeStringArray(br);
        }else if(elem == boolType){
            return DecodeBoolArray(br);
        }else if(elem == longType){
            return DecodeLongArray(br);
        }else{
            int len = br.ReadInt32();
            var _tmp = Array.CreateInstance(elem, len);
            for(int i = 0; i < len; i++){
                if(elem.IsArray){
                    _tmp.SetValue(DecodeArray(br, elem), i);
                }else{
                    _tmp.SetValue(DecodeStruct(br, elem), i);
                }
            }
            return _tmp;
        }
    }

    public static T DecodeDictionary<T>(Func<BinaryReader, T> func, BinaryReader br){
        return func(br);
    }

    public static T DecodeArray<T>(Func<BinaryReader, T> func, BinaryReader br){
        return func(br);
    }

    public static int[] DecodeIntArray(BinaryReader br){
        int len = br.ReadInt32();
        var _tmp = new int[len];
        for(int i = 0; i < len; i++){
            var item = br.ReadInt32();
            _tmp[i] = item;
        }
        return _tmp;
    }

    public static int[] DecodeInt8Array(BinaryReader br){
        int len = br.ReadInt32();
        var _tmp = new int[len];
        for(int i = 0; i < len; i++){
            var item = br.ReadSByte();
            _tmp[i] = item;
        }
        return _tmp;
    }

    public static int[] DecodeInt16Array(BinaryReader br){
        int len = br.ReadInt32();
        var _tmp = new int[len];
        for(int i = 0; i < len; i++){
            var item = br.ReadInt16();
            _tmp[i] = item;
        }
        return _tmp;
    }

    public static float[] DecodeLongArray(BinaryReader br){
        int len = br.ReadInt32();
        var _tmp = new float[len];
        for(int i = 0; i < len; i++){
            var item = br.ReadInt64();
            _tmp[i] = item;
        }
        return _tmp;
    }

    public static float[] DecodeFloatArray(BinaryReader br){
        int len = br.ReadInt32();
        var _tmp = new float[len];
        for(int i = 0; i < len; i++){
            var item = br.ReadSingle();
            _tmp[i] = item;
        }
        return _tmp;
    }

    public static bool[] DecodeBoolArray(BinaryReader br){
        int len = br.ReadInt32();
        var _tmp = new bool[len];
        for(int i = 0; i < len; i++){
            var item = br.ReadInt32() == 1 ? true : false;
            _tmp[i] = item;
        }
        return _tmp;
    }

    public static string[] DecodeStringArray(BinaryReader br){
        int len = br.ReadInt32();
        var _tmp = new string[len];
        for(int i = 0; i < len; i++){
            var item = System.Text.Encoding.UTF8.GetString(br.ReadBytes(br.ReadInt32()));;
            _tmp[i] = item;
        }
        return _tmp;
    }

    public static string DecodeString(BinaryReader br){
        int len = br.ReadInt32();
        if(len > 9999){
            Debug.LogError("解码字符串过长：" + len);
            return "";
        }
        br.Read(strBuffer, 0, len); 
        return System.Text.Encoding.UTF8.GetString(strBuffer, 0, len);
    }
}"""

class ValueTree(object):
    def __init__(self, parent, typeNode):
        self.leftVals = []
        self.rightVals = []
        self.parent = parent
        if self.parent != None:
            self.parent.Append(self)
        self.typeNode = typeNode
        self.isRight = False

    def Append(self, val):
        if self.isRight:
            self.rightVals.append(val)
        else:
            self.leftVals.append(val)

    def __str__(self):
        return (self.typeNode.ftype == "dict" and str(len(self.leftVals)) + "," + str(len(self.rightVals)) or str(len(self.leftVals))) + "<" + str(self.typeNode) + ">"
    
    def CheckValueValid(self):
        if self.typeNode.IsStruct():
            if len(self.leftVals) != len(self.typeNode.structTypes): #数组
                print "Warning: struct value is not valid, auto fix it", str(self)
                for i in range(len(self.leftVals), len(self.typeNode.structTypes)):
                    # print len(self.leftVals), len(self.typeNode.structTypes), i, self.typeNode.structTypes[i].GetDefaultValue()
                    self.leftVals.append(self.typeNode.structTypes[i].GetDefaultValue())
        elif self.typeNode.IsDictionary(): #字典的key value对不匹配
            if len(self.leftVals) > len(self.rightVals):
                print "Warning: dict value is not valid, auto fix it", str(self)
                for i in range(len(self.rightVals), len(self.leftVals)):
                    self.rightVals.append(self.typeNode.right.GetDefaultValue())
            elif len(self.leftVals) < len(self.rightVals):
                print "Warning: dict value is not valid, auto fix it", str(self)
                for i in range(len(self.rightVals), len(self.leftVals)):
                    self.rightVals.append(self.typeNode.right.GetDefaultValue())
        # elif self.typeNode.IsArray():
            

class TypeTree(object):
    def __init__(self, ftype, parent = None, left = None, right = None):
        self.ftype = ftype
        self.parent = parent
        self.left = left
        self.right = right

        self.fname = None
        self._typeIdx = 0

        if self.parent != None:
            if self.parent.ftype == "dict":
                if self.parent.left == None:
                    self.parent.left = self
                elif self.parent.right == None:
                    self.parent.right = self
                else:
                    print "Error: type is wrong", parent
            elif self.parent.ftype == "array":
                if self.parent.left != None:
                    print "Error: type is wrong2", parent
                self.parent.left = self
            # else:
            #     print "Error: type is wrong3", parent

        self._structTypeIdx = 0
        if self.ftype.startswith("&"):
            t = self.ftype[1:]
            if struct_info.has_key(t):
                tmpdef = struct_info[t]["define"]
                print tmpdef
                self.structTypes = []
                for k,v in tmpdef.iteritems():
                    tmp = ParseTypeDefine(v, self)# TypeTree(v, self,)
                    if tmp == None:
                        print "Erro: struct type {0} parse error".format(v)
                    tmp.fname = k
                    self.structTypes.append(tmp)
            else:
                print "Error: type {0} undefined ".format(t)
            
        # print self.name, self.parent
    
    def __str__(self):
        return self.ftype  + (self.left == None and "_" or ("<" + str(self.left)  + (self.ftype == "dict" and "," + str(self.right) or "")  + ">"))

    def CheckTypeIsComplete(self):
        return (self.ftype == "array" and self.left != None)  or \
               (self.ftype == "dict" and self.left != None and self.right != None) or \
               (not self.ftype.startswith("&") and self.ftype != "array" and self.ftype != "dict")

    def IsPrimitiveType(self):
        if self.ftype == "int32" or self.ftype == "int" or \
            self.ftype == "int8" or self.ftype == "int16" or \
            self.ftype == "int64" or self.ftype == "bool" or \
            self.ftype == "string" or self.ftype == "float":
            return True
        else:
            return False
    def IsStruct(self):
        return self.ftype.startswith("&")

    def IsArray(self):
        return self.ftype == "array"
        
    def IsDictionary(self):
        return self.ftype == "dict"

    def GetStructType(self, moveNext = True):
        if self.structTypes == None or len(self.structTypes) == 0:
            print "Error: struct type: {0} is wrong".format(self.ftype)
        else:
            tmp = self.structTypes[self._structTypeIdx]
            # print "struct type:", tmp
            if moveNext:
                self._structTypeIdx = (self._structTypeIdx + 1) % len(self.structTypes)
            return tmp

    def GetDefaultValue(self):
        if self.IsPrimitiveType():
            if self.ftype == "string":
                return ""
            else:
                return "0"
        elif self.IsStruct(): #结构体必须要有值
            val = ValueTree(None, self)
            for k in self.structTypes:
                val.Append(k.GetDefaultValue())
            return val
        else:#数组和字典可以是空的
            val = ValueTree(None, self)
            return val

    ##lenStr是传入的数组初始化长度
    def GetTypeName(self, arrLen = None):
        if self.ftype == "int32" or self.ftype == "int":
            return "int"
        elif self.ftype == "int8":
            return "sbyte"
        elif self.ftype == "int16":
            return "short"
        elif self.ftype == "int64":
            return "long"
        elif self.ftype == "bool":
            return "bool"
        elif self.ftype == "string":
            return "string"
        elif self.ftype == "float":
            return "float"
        elif self.IsStruct():
            t = self.ftype[1:]
            if struct_info.has_key(t):
                return struct_info[t]["class_name"]
            else:
                print u"无法识别的结构体", t
        elif self.IsArray():
            if arrLen == None:
                return self.left.GetTypeName() + "[]" 
            else:
                return self.left.IsArray() and self.left.GetTypeName(arrLen) + "[]" or self.left.GetTypeName() + "[" + arrLen + "]"
        elif self.IsDictionary():
            return "Dictionary<{},{}>".format(self.left.GetTypeName(), self.right.GetTypeName())
        else:
            print u"无法识别的数据类型1:", self.ftype
            # traceback.print_stack()
            return "string"   


    def GetReadFunc(self):
        if self.ftype == "int32" or self.ftype == "int":
            return u"br.ReadInt32();"
        elif self.ftype == "int8":
            return u"br.ReadSByte();"
        elif self.ftype == "int16":
            return u"br.ReadInt16();"
        elif self.ftype == "bool":
            return u"br.ReadInt32() == 1 ? true : false;"
        elif self.ftype == "string":
            return u"ConfigUtil.DecodeString(br);"
        elif self.ftype == "float":
            return u"br.ReadSingle();"
        elif self.ftype == "int64":
            return u"br.ReadInt64();"
        elif self.IsStruct():
            # t = typeStr[1:]
            return u"({0})ConfigUtil.DecodeStruct(br, typeof({0}));".format(self.GetTypeName())
        elif self.IsArray():
            # rand = random.randint(0,1000000000)
            rand = abs(hash(self.GetTypeName())) % (10 ** 8)
            # rand = self._typeIdx
            # self._typeIdx = self._typeIdx + 1
            return u"""ConfigUtil.DecodeArray<{0}>((reader_{2})=>{{
                int len_{2} = reader_{2}.ReadInt32();
                var arr_{2} = new {3};
                for(int i_{2} = 0; i_{2} < len_{2}; i_{2}++){{
                   arr_{2}[i_{2}] = {1}
                }}
                return arr_{2};
            }}, br);""".format(self.GetTypeName(),self.left.GetReadFunc(),rand,self.GetTypeName("len_{0}".format(rand)))
            # return "br.ReadInt();"
        elif self.IsDictionary():
            rand = abs(hash("Dictionary<{0},{1}>".format(self.left.GetTypeName(),self.right.GetTypeName()))) % (10 ** 8)
            # rand = random.randint(0,1000000000)
            # rand = self._typeIdx
            # self._typeIdx = self._typeIdx + 1
            return u"""ConfigUtil.DecodeDictionary<Dictionary<{0},{1}>>((reader_{4})=>{{
                int len_{4} = reader_{4}.ReadInt32();
                var dic_{4} = new Dictionary<{0},{1}>();
                for(int i_{4} = 0; i_{4} < len_{4}; i_{4}++){{
                   var key_{4} = {2}
                   var value_{4} = {3}
                   if(!dic_{4}.ContainsKey(key_{4})) dic_{4}.Add(key_{4},value_{4}); else Debug.LogError("重复的key：" + key_{4});
                }}
                return dic_{4};
            }}, br);
            """.format(self.left.GetTypeName(), self.right.GetTypeName(), self.left.GetReadFunc(), self.right.GetReadFunc(), rand)
        else:
            print u"获取函数失败，无法识别的数据类型：", self.ftype
            # traceback.print_stack()
            return u"br.ReadInt();"  

def ParseTypeDefine(etype, parent = None):
    # parent = None
    root = None
    for k in list(filter(None,re.split( r"[ ,<>]", etype))):
        if k == "array":
            newType = TypeTree(k, parent)
            parent = newType
        elif k == "dict":
            newType = TypeTree(k, parent)
            parent = newType
        elif k.startswith("&"):
            newType = TypeTree(k, parent)
            #因为一个类型解析完毕之后，还需要向上迭代检查父级类型是否也结束了
            while(parent != None and parent.parent != None and parent.CheckTypeIsComplete()): 
                parent = parent.parent
        else: #elif 
            newType = TypeTree(k, parent)
            while(parent != None and parent.parent != None and parent.CheckTypeIsComplete()): 
                parent = parent.parent
        if root == None: root = newType

    return root

def ParseToken(content, etype):
    if etype.IsPrimitiveType():
        return content

    cnt = len(content)
    i = 0
    QuoteMode = False
    EscapeMode = False

    root = None
    curVal = None
    currType = etype
    primitiveIdx = -1
    while i < cnt:
        token = content[i]

        if QuoteMode: #QuoteMode内才用转义字符
            if token == "\\":
                EscapeMode = True
                i = i + 1
                continue

        if EscapeMode:
            if token == "\"" or token == "\\":
                EscapeMode = False
                i = i + 1
                continue
            else:
                print "Error: escape character in wrong place", i, evalue
                EscapeMode = False
                i = i + 1
                continue

        if token == "\"":
            QuoteMode = not QuoteMode
            if not QuoteMode: #字符串结束
                # curVal.Append(evalue[primitiveIdx:i])
                print evalue[primitiveIdx:i]
                i = i + 1
                continue

        if QuoteMode:
            if currType.ftype == "string" and primitiveIdx == -1:
                primitiveIdx = i
            i = i + 1
            continue

        if token == "[":
            primitiveIdx = -1
            curVal = ValueTree(curVal, currType)
            currType = currType.left #数组元素类型
        elif token == "<":
            primitiveIdx = -1
            curVal = ValueTree(curVal, currType)
            currType = currType.left
        elif token == "{" or token == "(":
            primitiveIdx = -1
            if not currType.IsStruct():
                print "Error: value conflicted with type", currType, i, evalue
            else:
                curVal = ValueTree(curVal, currType)
                currType = currType.GetStructType() #进入到结构体的第一个字段
        elif token == "]": #数组结束
            if currType.IsPrimitiveType():
                if primitiveIdx == -1:
                    # print "Error: value format not right", currType, i, evalue
                    curVal.Append("")
                else:
                    curVal.Append(evalue[primitiveIdx:i])
            curVal.CheckValueValid()
            curVal = curVal.parent
            currType = currType.parent
            primitiveIdx = -1
        elif token == "}" or token == ")": #结构体结束
            if currType.IsPrimitiveType():
                if primitiveIdx == -1:
                    print "Error: value format not right", currType, i, evalue
                else:
                    curVal.Append(evalue[primitiveIdx:i])
            curVal.CheckValueValid()
            curVal = curVal.parent
            currType = currType.parent
            primitiveIdx = -1

            # print "struct back type:", currType
        elif token == ">": #字典结束
            if currType.IsPrimitiveType():
                if primitiveIdx == -1:
                    print "Error: value format not right", currType, i, evalue
                    curVal.Append("")
                else:
                    curVal.Append(evalue[primitiveIdx:i])
            curVal.CheckValueValid()
            curVal = curVal.parent
            currType = currType.parent
            primitiveIdx = -1
            
        elif token == ":": #冒号分隔 example: <key:value, key1:value1>
            if currType.parent.IsDictionary():
                if currType.IsPrimitiveType():
                    if primitiveIdx == -1:
                        print "Error: value format not right", currType, i, evalue
                        curVal.Append("")
                    else:
                        curVal.Append(evalue[primitiveIdx:i])
                currType = currType.parent.right
                curVal.isRight = True
                # curVal = ValueTree(curVal, currType)
            else:
                print "Error: value format not right, \":\" can only be in dictionary", currType, evalue
            primitiveIdx = -1
            
        elif token == ",":
            if currType.IsPrimitiveType():
                if primitiveIdx == -1: ##连续两个逗号
                    val = ""
                else:
                    val = evalue[primitiveIdx:i]
                if currType.parent.IsDictionary(): #key,value的切换
                    curVal.Append(val)
                    curVal.isRight = False
                    currType = currType.parent.left
                    primitiveIdx = -1
                elif currType.parent.IsArray(): #数组项的增加
                    curVal.Append(val)
                    primitiveIdx = -1
                elif currType.parent.IsStruct(): #父级是结构体，逗号表示结构体字段的迁移
                    currType = currType.parent.GetStructType()
                    curVal.Append(val)
                    primitiveIdx = -1

            else:#逗号分隔有三种情况，struct,array,dictionary
                if currType.parent != None:
                    primitiveIdx = -1
                    if currType.parent.IsStruct(): #父级是结构体，逗号表示结构体字段的迁移
                        currType = currType.parent.GetStructType()
                    elif currType.parent.IsDictionary(): #开始另外一个key:value对
                        currType = currType.parent.left
                        curVal.isRight = False
                    elif currType.IsArray():
                        i = i + 1
                        continue
                else:
                    primitiveIdx = -1
                    print "Error: value format not right, comma in wrong place", evalue
                # if currType.IsStruct(): #这种情况是non primitive类型的数据结束之后会回退到struct
                # if currType.GetStructType(False).IsPrimitiveType(): #primitive类型因为没有显示的开始值的地方，所以如果是结构体，primitive类型需要显式移动字段
                #     currType = currType.GetStructType()
                
        elif token == " ": #quote mode以外的空格全部没有意义
            i = i + 1
            continue
        else:
            if currType.IsPrimitiveType() and primitiveIdx == -1:
                primitiveIdx = i

        print "cur type:{0} token:\"{1}\" val:{2}".format(currType, token, curVal)
                
        if root == None: root = curVal
        i = i + 1
    if root == None:
        print "Error: value is wrong", content
    return root

class CellValue(object):

    _type = None,
    _value = None,

    def __init__(self, eType, value):
        self._type = eType
        self._value = value

    @property
    def eType(self):
        return self._type

    @property
    def value(self):
        return self._value

class ConfigEntry(object):

    def __init__(self, row):
        self._row = row
        self._valueList = {}

    def AddCell(self, col, value):
        self._valueList[col] = value
        # pass

class ConfigTable(object):

    # _name = None
    # _typeList = [] #输出的字段类型列表
    # _colList = [] #输出的行号列表
    # _descList = [] #输出的字段描述列表
    # _fieldNameList = [] #输出的字段名字列表

    # _entryList = {} #{row,一行数据}的表

    # _startRow = 0
    # _endRow = 0
    # _startCol = 0
    # _endCol = 0

    def __init__(self, _name, colList, typeList, descList, fielNameList):
        self._name = _name
        self._typeList = typeList
        self._colList = colList
        self._descList = descList
        self._fieldNameList = fielNameList
        self._entryList = {}

        for i in range(0, len(self._fieldNameList)): #字段名
            self._typeList[i].fname = self._fieldNameList[i]
    
    def AddEntry(self, row, value):
        #self._typeList[idx]
        # if not self._entryList.__contains__(row):
        #     self._entryList[row] = ConfigEntry(row)
        # idx = self._colList.index(col)
        # self._entryList[row].AddCell(self._fieldNameList[idx], value)
        self._entryList[row] = value
        # print self._entryList[row]
        # pass

    def RemoveEntry(self, row):
        self._entryList[row] = None


def CheckDataType(typeStr):
    if typeStr == "int32" or\
       typeStr == "int64" or\
       typeStr == "int" or\
       typeStr == "bool" or\
       typeStr == "string" or\
       typeStr == "array" or\
       typeStr == "int8" or\
       typeStr == "int16" or\
       typeStr == "float":
       return 1
    elif typeStr.startswith("array<"):
        # if "<" in typeStr[6:]: #检测不能有数组嵌套
        #     return 0
        # else:
        return 1
    elif typeStr.startswith("dict<"):
        if "," not in typeStr:
            return 0
        else:
            return 1
    elif typeStr.startswith("&") and struct_info.has_key(typeStr[1:]):
        return 1
    else:
        return 0


def OutputUtilFile():
    i = 0
    funcBody = ""
    for key, val in struct_info.iteritems(): 
        if val.has_key("is_generate") and val["is_generate"] == 1:
            fields = ""
            for fieldName, fieldType in val["define"].iteritems():
                # fieldFormat.format()
                fields += fieldFormat.replace("@@fieldComment", "no comment").replace("@@fieldName", fieldName).replace("@@fieldType", ParseTypeDefine(fieldType).GetTypeName())
            struct = structFormat.replace("@@fields", fields).replace("@@className", val["class_name"])
            ofile = codecs.open(outputDir + val["class_name"] + ".cs", 'w', "utf-8")
            try:
                ofile.write(struct)
            except IOError, e:
                print u"ConfigUtil.cs 写入文件异常", e
            finally:
                ofile.close()
        if i == 0:
            funcBody += "        if(t == typeof({0})){{\n".format(val["class_name"])
        else:
            funcBody += "        }}else if(t == typeof({0})){{\n".format(val["class_name"])

        funcBody += "            var {0} = new {1}();\n".format(key, val["class_name"])
        
        # print type(val["define"])
        for fieldName, fieldType in val["define"].iteritems():
            tr = ParseTypeDefine(fieldType, None)
            funcBody += "            {0}.{1} = {2}\n".format(key, fieldName, tr.GetReadFunc())
        
        funcBody += "            return {0};\n".format(key)
        i = i + 1
    funcBody += "        }\n"
    func = decodeStructFunc.replace("@@funcBody", funcBody)
    util = configUtil.replace("@@DecodeStruct", func)

    ofile = codecs.open(outputDir + "ConfigUtil.cs", 'w', "utf-8")
    try:
        ofile.write(util)
    except IOError, e:
        print u"ConfigUtil.cs 写入文件异常", e
    finally:
        ofile.close()

def EncodeJsonData(ofile, typeTree, evalue):
    # if debug_info: print "encode value:", etype, type(evalue)
    if typeTree.IsPrimitiveType():
        if IsEmptyOrNull(evalue) : 
            if typeTree.ftype == "string": 
                evalue = ""
            else:
                evalue = "0"
        # evalue = unicode(evalue)
        if typeTree.ftype == "int" or typeTree.ftype == "int8" or typeTree.ftype == "int16" or typeTree.ftype == "int32":
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":{1}".format(typeTree.fname, evalue))
            else: ofile.write("{0}".format(evalue))
        elif typeTree.ftype == "bool":
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":{1}".format(typeTree.fname, evalue))
            else: ofile.write("{0}".format(evalue))
        elif typeTree.ftype == "int64":
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":{1}".format(typeTree.fname, evalue))         
            else: ofile.write("{0}".format(evalue))
        elif typeTree.ftype == "string":
            # print "unicode len", len(evalue.encode('utf-8')), evalue
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":\"{1}\"".format(typeTree.fname, evalue))
            else: ofile.write("\"{0}\"".format(evalue))
        elif typeTree.ftype == "float":
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":{1}".format(typeTree.fname, evalue))        
            else: ofile.write("{0}".format(evalue))
    else:
        if evalue == None or evalue.typeNode != typeTree:
            print "Error: typeTree:{0} is mismatch with typeValue:{1}".format(typeTree, evalue), traceback.extract_stack()
        if typeTree.IsStruct():
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":{{".format(typeTree.fname))
            else: ofile.write("{")

            i = 0
            cnt = len(evalue.leftVals)
            # print "value:", evalue.leftVals
            for v in evalue.leftVals:
                EncodeJsonData(ofile, typeTree.structTypes[i], v)
                if i < cnt - 1: ofile.write(",")
                i = i + 1

            ofile.write("}")
            
        elif typeTree.IsArray():##数组直接用逗号隔开
            
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":[".format(typeTree.fname))
            else: ofile.write("[")
            i = 0
            cnt = len(evalue.leftVals)
            for m in evalue.leftVals:
                EncodeJsonData(ofile, typeTree.left, m)
                if i < cnt - 1: ofile.write(",")
                i = i + 1
            ofile.write("]")

        elif typeTree.IsDictionary():
            if not IsEmptyOrNull(typeTree.fname): ofile.write("\"{0}\":{{".format(typeTree.fname))
            else: ofile.write("{")
            i = 0
            cnt = len(evalue.leftVals)
            for k in evalue.leftVals:
                EncodeJsonData(ofile, typeTree.left, k)
                ofile.write(":")
                EncodeJsonData(ofile, typeTree.right, evalue.rightVals[i])
                if i < cnt - 1: ofile.write(",")
                i = i + 1
            ofile.write("}")

        elif typeTree.ftype.startswith("script"):
            print u"字段是脚本读取，忽略"
        else:
            print "Error: encode data with unknown evalue", evalue

def IsEmptyOrNull(evalue):
    return evalue == "" or evalue == "None" or evalue == None


def EncodeData(ofile, typeTree, evalue):
    # if debug_info: print "encode value:", etype, type(evalue)
    if typeTree.IsPrimitiveType():
        if IsEmptyOrNull(evalue) : 
            if typeTree.ftype == "string": 
                evalue = ""
            else:
                evalue = "0"
        # evalue = unicode(evalue)
        # print "value:", evalue
        if typeTree.ftype == "int" or typeTree.ftype == "int8" or typeTree.ftype == "int16" or typeTree.ftype == "int32":
            ft = ""
            if typeTree.ftype == "int" or typeTree.ftype == "int32": ft = '<i' 
            elif typeTree.ftype == "int8": ft = '<b'
            elif typeTree.ftype == "int16": ft = '<h'
            # print "ft:",ft
            if "." in evalue:
                ofile.write(struct.pack(ft, int(evalue.rsplit(".")[0].encode("utf-8"))))
                print "Error: value type:{0} is conflicted with value:{1}".format(typeTree.ftype, evalue) 
            else:
                # try:
                ofile.write(struct.pack(ft, int(evalue.encode("utf-8"))))
                # except:
                    # print "encode value:", etype, evalue
        elif typeTree.ftype == "bool":
            ofile.write(struct.pack('<i', int(evalue.encode("utf-8"))))
        elif typeTree.ftype == "int64":
            if "." in evalue:
                print "Error: value type:{0} is conflicted with value:{1}".format(typeTree.ftype, evalue) 
                ofile.write(struct.pack('<q', int(evalue.rsplit(".")[0].encode("utf-8"))))
            else:
                ofile.write(struct.pack('<q', long(evalue.encode("utf-8"))))
        elif typeTree.ftype == "string":
            # print "unicode len", len(evalue.encode('utf-8')), evalue
            cnt = len(evalue.encode('utf-8'))
            ofile.write(struct.pack('<i', cnt))
            content = struct.pack('<'+ str(cnt) + 's',  unicode(evalue).encode("utf-8"))
            # print content.encode("utf-8")
            ofile.write(content)
        elif typeTree.ftype == "float":
            ofile.write(struct.pack('<f', float(evalue.encode("utf-8"))))
    else:
        if evalue.typeNode != typeTree:
            print "Error: typeTree:{0} is mismatch with typeValue:{1}".format(typeTree, evalue)
        if typeTree.IsStruct():
            i = 0
            for v in evalue.leftVals:
                EncodeData(ofile, typeTree.structTypes[i], v)
                i = i + 1
            
        elif typeTree.IsArray():##数组直接用逗号隔开
            ofile.write(struct.pack('<i', len(evalue.leftVals))) #写入数组长度
            for v in evalue.leftVals:
                EncodeData(ofile, typeTree.left, v)
        elif typeTree.IsDictionary():
            ofile.write(struct.pack('<i', len(evalue.leftVals))) #写入keys长度
            i = 0
            for k in evalue.leftVals:
                EncodeData(ofile, typeTree.left, k)
                EncodeData(ofile, typeTree.right, evalue.rightVals[i])
                i = i + 1
                
        elif typeTree.startswith("script"):
            print u"字段是脚本读取，忽略"
        else:
            print "Error: encode data with unknown evalue", evalue

def GetSheetClassName(sheetName):
    ret = ""
    for s in sheetName[5:].split("_"):
        ret = ret + s[0].upper() + s[1:]
    return "Entity" + ret

fn = sys.argv[1]
try:
    wb = load_workbook(fn, read_only=True,data_only=True)
    sheetnames = wb.sheetnames
    sheet = wb[sheetnames[0]]
    # color = sheet.sheet_properties.tabColor
    if sheet.title.startswith("data_"):
        if True:#color.rgb == "FF92D050" or color.rgb == "FF00B050": 
            startRow = 0
            endRow = 0
            startCol = 0
            endCol = 0
            # for row in sheet.rows:
            print u"\n开始处理表：" + sheet.title + " ---max row: " + str(sheet.max_row) + " ---max colum: " + str(sheet.max_column)
            rowIdx = 1

            #在表中寻找起始点
            for row in sheet.iter_rows():
                colIdx = 1
                for cell in row:
                    val = unicode(cell.value)
                    if val.find("!!") >= 0:## val == "!!cs" or val == "!!sc" or val == "!!s" or val == "!!c":
                        startRow = cell.row
                        startCol = cell.column
                        break
                if startRow > 0: #找到了起始点
                    print "start point found"
                    break

            # print "output data"
            #获得所有的导出列

            # table = {}
            if startRow > 0:
                colIdx = 0
                clientOut = []
                serverOut = []

                outputType = []
                outputDesc = []
                outputFileName = []
                # while(colIdx <= sheet.max_column):
                curRow = startRow
                for row in sheet.iter_rows(None, startRow, startRow + 4):
                    colIdx = 0
                    for cell in row:
                    # cell = sheet.cell(startRow, colIdx)
                        colIdx = colIdx + 1
                        if curRow == startRow: ##第一行是标记客户端和服务器的输出字段
                            if cell.value is not None:
                                val = unicode(cell.value)
                                if "c" in val:## or "sc" in val or "!!cs" in val or "!!sc" in val:
                                    clientOut.append(cell.column)
                                    serverOut.append(cell.column)
                                # elif val == "c" or val == "!!c":
                                #     clientOut.append(cell.column)
                                elif "s" in val:## == "!!s":
                                    serverOut.append(cell.column)
                        elif curRow == startRow + 1: #字段类型
                            if colIdx in clientOut:
                                if cell.value is None:
                                    print u"Excel表格式有误：请在c/s的下一行标注数据类型，列号：" , unicode(colIdx)
                                    outputType.append(ParseTypeDefine("string"))
                                else:
                                    if CheckDataType(cell.value):
                                        outputType.append(ParseTypeDefine(cell.value))
                                    else:
                                        print u"Excel表格式有误：标注的数据类型无法解析->" , unicode(cell.value)
                                        outputType.append(ParseTypeDefine("string")) #默认输出为string
                        elif curRow == startRow + 2: #字段描述
                            if colIdx in clientOut:
                                if cell.value is None:
                                    print u"Excel表格式有误：请在c/s的下面第二行标注字段描述，列号：" , unicode(colIdx)
                                    outputDesc.append("暂无描述")
                                else:
                                    outputDesc.append(unicode(cell.value))
                        elif curRow == startRow + 3: #字段名
                            if colIdx in clientOut:
                                print u"字段：", colIdx, unicode(cell.value)
                                if cell.value is None:
                                    print u"Excel表格式有误：请在c/s的下面第三行标注字段名称，列号：" , unicode(colIdx)
                                    outputFileName.append(unicode("field_name"))
                                else:
                                    outputFileName.append(unicode(cell.value))   
                    curRow = curRow + 1
                    # colIdx = colIdx + 1
            else:
                print u"没有标记导出起始位置，导出失败：",sheet.title
            print u"客户端导出列：", clientOut, u"服务器导出列：", serverOut
                
            print u"读入数据"
            #初始化表
            table = ConfigTable(sheet.title, clientOut, outputType, outputDesc, outputFileName)

            #开始读取数据
            dataRow = startRow + 4
            for row in sheet.iter_rows(None, dataRow):
                blankCell = 0
                entry = ConfigEntry(dataRow)
                curCol = 0
                for cell in row:
                    curCol = curCol + 1
                    if  curCol in clientOut:
                        if cell.value is not None and not isinstance(cell, EmptyCell):
                            #直接添加索引值，index()没有找到会抛出错误，所以先判断
                            cooolumn = clientOut.index(cell.column) if cell.column in clientOut else -1
                            # print coool, cell.value
                            # if cooolumn >= 0: outputType[cooolumn]
                            entry.AddCell(cooolumn, cell.value)
                        else:
                            entry.AddCell(clientOut.index(curCol) if curCol in clientOut else -1, None)
                            if curCol == clientOut[0]:
                                blankCell = 1
                            # if blankCell > 3:
                            #     break
                            # if clientOut.index(curCol) > 3:
                                # print u"Excel表数据有误：导出的字段值不能为空，行：", unicode(dataRow) , u" 列：" ,unicode(curCol)
                if blankCell > 0:
                    print u"初始列的值为空，到达数据终点，行：", dataRow
                    break
                else:
                    table.AddEntry(dataRow, entry)
                    # print "entry:",dataRow, entry._valueList
                    # if dataRow > startRow + 5:
                    #     print "entry1:",dataRow, table._entryList[dataRow-1]._valueList
                dataRow = dataRow + 1
            
#-----------------------Generate Class File Start
            print u"输出CSharp类文件", table._fieldNameList
            #保存文件格式
            fields = ""
            decodeFunc = ""
            prefix = "    " 
            idx = 0
            hasLen = False
            for f in table._fieldNameList:
                print "analyze field:", f
                tr = table._typeList[idx]
                if f != "ID": ##Entity父类有这个属性
                    fields = fields + fieldFormat.replace("@@fieldComment", table._descList[idx]).replace("@@fieldName", f).replace("@@fieldType", tr.GetTypeName())
                decodeFunc = decodeFunc + prefix + prefix + u"this.{0} = {1}\n".format(f, tr.GetReadFunc())
                    
                idx = idx + 1

            clsName = GetSheetClassName(sheet.title)
            classStr = classFormat.replace("@@className", clsName).replace("@@fields", fields)
            classStr = classStr.replace("@@funcBody", decodeFunc)
            # print decodeFunc

            ofile = codecs.open(outputDir + clsName + ".cs", 'w', "utf-8")
            try:
                ofile.write(classStr)
            except IOError,e:
                print u"写入文件异常", clsName, e
            finally:
                ofile.close()
#-----------------------Generate Class File End


#-----------------------Output bytes File Start
            ofile = file(dataDir + sheet.title + ".bytes", "wb+", 1)
            jsonFile = file(dataDir + sheet.title + ".json", "wb+", 1)

            try:
                cnt1 = len(table._entryList)
                ofile.write(struct.pack('<i', cnt1))
                print u"数据总长度:", cnt1
                jsonFile.write('[')

                j = 0
                for dval in table._entryList.itervalues():
                    cnt = len(dval._valueList)
                    jsonFile.write("{")

                    i = 0
                    for ekey, evalue in dval._valueList.iteritems(): #entry: {colIdx, string}
                        # print "key-value:", ekey, evalue
                        evalue = unicode(evalue)
                        if ekey < 0:
                            print u"导出数据错误：数据不在导出列表中->", evalue
                        else:
                            tr = table._typeList[i]
                            #修正数组的配置
                            if tr.IsArray() and not evalue.startswith("["): evalue = "[{0}]".format(evalue)
                            if tr.IsStruct() and not evalue.startswith("{"): evalue = "{{{0}}}".format(evalue)
                            val = ParseToken(evalue, tr)
                            
                            EncodeJsonData(jsonFile, tr, val)
                            if i < cnt - 1: jsonFile.write(",")
                            EncodeData(ofile, tr, val)
                                # print u"无法解析的数据类型", etype
                        i = i + 1
                    if j < cnt1 - 1: jsonFile.write("},")
                    else: jsonFile.write("}")
                    j = j + 1
                jsonFile.write("]")
            except IOError, e:
                print u"写入文件异常:", sheet.title, str(e)
            finally:
                ofile.close()
    wb.close()
#-----------------------Output bytes File End

    OutputUtilFile()

except IOError, e:
    print "read file error: " + fn + " " + str(e)


    # for col in sheet.columns:
    #     for c in col:
    #         print c.value
#       
time.sleep(1)

# os.system("pause")